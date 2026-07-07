# Copyright (c) 2010-2024 fastpyxl

from __future__ import annotations

import datetime
from dataclasses import dataclass
from pathlib import Path
from typing import Any

@dataclass(frozen=True)
class CellSnapshot:
    value: Any
    data_type: str | None
    number_format: str | None


@dataclass(frozen=True)
class SheetSnapshot:
    title: str
    max_row: int
    max_column: int
    cells: dict[str, CellSnapshot]


@dataclass(frozen=True)
class WorkbookSnapshot:
    sheetnames: tuple[str, ...]
    active_title: str
    sheets: dict[str, SheetSnapshot]


def _normalize_value(value: Any) -> Any:
    if isinstance(value, datetime.datetime):
        return value.replace(tzinfo=None)
    if isinstance(value, datetime.time):
        return value.replace(tzinfo=None)
    return value


def snapshot_workbook(workbook: Any) -> WorkbookSnapshot:
    sheets: dict[str, SheetSnapshot] = {}
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        cells: dict[str, CellSnapshot] = {}
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cells[cell.coordinate] = CellSnapshot(
                    value=_normalize_value(cell.value),
                    data_type=getattr(cell, "data_type", None),
                    number_format=getattr(cell, "number_format", None),
                )
        sheets[name] = SheetSnapshot(
            title=worksheet.title,
            max_row=worksheet.max_row,
            max_column=worksheet.max_column,
            cells=cells,
        )
    return WorkbookSnapshot(
        sheetnames=tuple(workbook.sheetnames),
        active_title=workbook.active.title,
        sheets=sheets,
    )


def assert_workbook_snapshots_equal(
    fast_snapshot: WorkbookSnapshot,
    openpyxl_snapshot: WorkbookSnapshot,
) -> None:
    assert fast_snapshot.sheetnames == openpyxl_snapshot.sheetnames
    assert fast_snapshot.active_title == openpyxl_snapshot.active_title
    assert set(fast_snapshot.sheets) == set(openpyxl_snapshot.sheets)
    for sheet_name in fast_snapshot.sheetnames:
        fast_sheet = fast_snapshot.sheets[sheet_name]
        openpyxl_sheet = openpyxl_snapshot.sheets[sheet_name]
        assert fast_sheet.title == openpyxl_sheet.title
        assert fast_sheet.max_row == openpyxl_sheet.max_row
        assert fast_sheet.max_column == openpyxl_sheet.max_column
        assert set(fast_sheet.cells) == set(openpyxl_sheet.cells)
        for coordinate, fast_cell in fast_sheet.cells.items():
            openpyxl_cell = openpyxl_sheet.cells[coordinate]
            assert fast_cell.value == openpyxl_cell.value, (
                f"{sheet_name}!{coordinate} value mismatch: "
                f"fastpyxl={fast_cell.value!r} openpyxl={openpyxl_cell.value!r}"
            )
            assert fast_cell.data_type == openpyxl_cell.data_type, (
                f"{sheet_name}!{coordinate} data_type mismatch: "
                f"fastpyxl={fast_cell.data_type!r} openpyxl={openpyxl_cell.data_type!r}"
            )
            assert fast_cell.number_format == openpyxl_cell.number_format, (
                f"{sheet_name}!{coordinate} number_format mismatch: "
                f"fastpyxl={fast_cell.number_format!r} "
                f"openpyxl={openpyxl_cell.number_format!r}"
            )


def assert_workbooks_match(fast_workbook: Any, openpyxl_workbook: Any) -> None:
    assert_workbook_snapshots_equal(
        snapshot_workbook(fast_workbook),
        snapshot_workbook(openpyxl_workbook),
    )


def load_fixture_both(path: Path, **kwargs: Any) -> tuple[Any, Any]:
    import fastpyxl
    import openpyxl

    return (
        fastpyxl.load_workbook(path, **kwargs),
        openpyxl.load_workbook(path, **kwargs),
    )


def save_and_reload_cross(
    writer_workbook: Any,
    reader_load: Any,
    path: Path,
) -> Any:
    writer_workbook.save(path)
    return reader_load(path)
