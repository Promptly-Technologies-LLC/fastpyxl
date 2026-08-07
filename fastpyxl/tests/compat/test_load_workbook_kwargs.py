# Copyright (c) 2010-2024 fastpyxl

from __future__ import annotations

from pathlib import Path

import pytest

from .conftest import (
    EXTERNAL_LINKS_FIXTURE,
    GENUINE_SAMPLE_FIXTURE,
    LEGACY_DRAWING_XLSM,
    READ_FIXTURE_IDS,
    READ_FIXTURE_PATHS,
)
from .helpers import assert_workbooks_match, load_fixture_both


pytestmark = pytest.mark.openpyxl_compat


def _close_pair(fast_wb, openpyxl_wb) -> None:
    if hasattr(fast_wb, "close"):
        fast_wb.close()
    if hasattr(openpyxl_wb, "close"):
        openpyxl_wb.close()


@pytest.mark.parametrize("data_only", [False, True])
def test_data_only_load_matches(data_only: bool):
    fast_wb, openpyxl_wb = load_fixture_both(
        GENUINE_SAMPLE_FIXTURE, data_only=data_only
    )
    try:
        assert fast_wb.data_only is data_only
        assert openpyxl_wb.data_only is data_only
        assert_workbooks_match(fast_wb, openpyxl_wb)
        formula_sheet = "Sheet3 - Formulas"
        fast_value = fast_wb[formula_sheet]["D2"].value
        openpyxl_value = openpyxl_wb[formula_sheet]["D2"].value
        if data_only:
            assert fast_value == 5
            assert openpyxl_value == 5
        else:
            assert fast_value == "='Sheet2 - Numbers'!D5"
            assert openpyxl_value == "='Sheet2 - Numbers'!D5"
    finally:
        _close_pair(fast_wb, openpyxl_wb)


@pytest.mark.parametrize("keep_vba", [False, True])
def test_keep_vba_cell_values_and_presence_match(keep_vba: bool):
    fast_wb, openpyxl_wb = load_fixture_both(LEGACY_DRAWING_XLSM, keep_vba=keep_vba)
    try:
        assert_workbooks_match(fast_wb, openpyxl_wb)
        assert (fast_wb.vba_archive is None) is (not keep_vba)
        assert (openpyxl_wb.vba_archive is None) is (not keep_vba)
    finally:
        _close_pair(fast_wb, openpyxl_wb)


@pytest.mark.xfail(
    reason=(
        "Intentional divergence: fastpyxl keeps only VBA-relevant parts in "
        "vba_archive, while openpyxl 3.1.5 stores the full package archive."
    ),
    strict=True,
)
def test_keep_vba_archive_namelist_matches_openpyxl():
    fast_wb, openpyxl_wb = load_fixture_both(LEGACY_DRAWING_XLSM, keep_vba=True)
    try:
        assert fast_wb.vba_archive is not None
        assert openpyxl_wb.vba_archive is not None
        assert sorted(fast_wb.vba_archive.namelist()) == sorted(
            openpyxl_wb.vba_archive.namelist()
        )
    finally:
        _close_pair(fast_wb, openpyxl_wb)


@pytest.mark.parametrize("keep_links", [False, True])
def test_keep_links_external_links_match(keep_links: bool):
    fast_wb, openpyxl_wb = load_fixture_both(
        EXTERNAL_LINKS_FIXTURE, keep_links=keep_links
    )
    try:
        assert_workbooks_match(fast_wb, openpyxl_wb)
        assert len(fast_wb._external_links) == len(openpyxl_wb._external_links)
        if keep_links:
            assert len(fast_wb._external_links) == 1
        else:
            assert fast_wb._external_links == []
            assert openpyxl_wb._external_links == []
    finally:
        _close_pair(fast_wb, openpyxl_wb)


@pytest.fixture
def rich_text_fixture(tmp_path: Path) -> Path:
    import openpyxl
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont

    path = tmp_path / "rich_text.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rich"
    ws["A1"] = CellRichText(
        TextBlock(InlineFont(b=True, color="FF0000"), "Hello"),
        " world",
    )
    ws["B1"] = "plain"
    wb.save(path)
    return path


@pytest.mark.parametrize("rich_text", [False, True])
def test_rich_text_load_matches(rich_text: bool, rich_text_fixture: Path):
    fast_wb, openpyxl_wb = load_fixture_both(rich_text_fixture, rich_text=rich_text)
    try:
        assert_workbooks_match(fast_wb, openpyxl_wb)
        fast_value = fast_wb["Rich"]["A1"].value
        openpyxl_value = openpyxl_wb["Rich"]["A1"].value
        if rich_text:
            assert type(fast_value).__name__ == "CellRichText"
            assert type(openpyxl_value).__name__ == "CellRichText"
            assert str(fast_value) == str(openpyxl_value) == "Hello world"
        else:
            assert fast_value == openpyxl_value == "Hello world"
            assert isinstance(fast_value, str)
            assert isinstance(openpyxl_value, str)
    finally:
        _close_pair(fast_wb, openpyxl_wb)


@pytest.mark.parametrize("fixture_path", READ_FIXTURE_PATHS, ids=READ_FIXTURE_IDS)
def test_read_only_load_matches_across_fixtures(fixture_path: Path):
    fast_wb, openpyxl_wb = load_fixture_both(fixture_path, read_only=True)
    try:
        assert fast_wb.read_only is True
        assert openpyxl_wb.read_only is True
        assert_workbooks_match(fast_wb, openpyxl_wb)
    finally:
        _close_pair(fast_wb, openpyxl_wb)
